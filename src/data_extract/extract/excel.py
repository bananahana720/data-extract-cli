"""Excel extractor adapter."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Tuple

from data_extract.extract.adapter import ExtractorAdapter


class ExcelExtractorAdapter(ExtractorAdapter):
    """Extractor for Excel workbooks."""

    def __init__(self) -> None:
        super().__init__(format_name="excel")

    def extract(self, file_path: Path) -> Tuple[str, Dict[str, Any], Dict[str, float]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel extraction") from exc

        workbook = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        sheet_lines = []
        row_count = 0
        max_cols = 0

        for sheet in workbook.worksheets:
            sheet_lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_values = ["" if cell is None else str(cell) for cell in row]
                while row_values and row_values[-1] == "":
                    row_values.pop()
                if not row_values:
                    continue
                sheet_lines.append(" | ".join(row_values))
                row_count += 1
                max_cols = max(max_cols, len(row_values))

        text = "\n".join(sheet_lines)
        structure = {
            "sheet_count": len(workbook.worksheets),
            "row_count": row_count,
            "column_count": max_cols,
        }
        quality = {
            "extraction_confidence": 1.0,
        }
        return text, structure, quality
